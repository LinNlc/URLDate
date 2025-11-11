from __future__ import annotations

import shutil
from pathlib import Path
from typing import Callable, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..utils.logger import get_logger
from . import image_fetcher
from .config_store import load_mode
from .staff_db import load_staff_database, match_staff_id
from .text_utils import check_content_length, has_invalid_actor_delimiter, is_valid_url

LoggerCallback = Callable[[str, str], None]
ProgressCallback = Callable[[int, int, str], None]

BASE_DIR = Path(__file__).resolve().parents[2]
TEMP_DIR = BASE_DIR / "temp"


class ExcelProcessor:
    def __init__(
        self,
        logger_callback: Optional[LoggerCallback] = None,
        progress_callback: Optional[ProgressCallback] = None,
    ) -> None:
        self.logger = get_logger(__name__)
        if logger_callback:
            def emitter(message: str, level: str = "info") -> None:
                logger_callback(message, level)
                log_method = getattr(self.logger, level, self.logger.info)
                log_method(message)
            self._emit_log = emitter
        else:
            self._emit_log = lambda message, level="info": getattr(self.logger, level, self.logger.info)(message)
        self._emit_progress = progress_callback or (lambda current, total, desc: None)
        self._temp_files: List[Path] = []

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def process(self, file_path: str, mode: Optional[int] = None) -> List[Path]:
        mode = mode or load_mode()
        try:
            workbook = self._load_workbook(file_path)
            base_name = Path(file_path).with_suffix("")
            sheet = workbook.active
            if sheet is None:
                raise ValueError("无法获取活动工作表")

            data_row_count = max(sheet.max_row - 1, 0)
            if mode == 1 and data_row_count > 50:
                self._emit_log("数据行数超过50，按模式1拆分为多个文件...", "warning")
                output_files = self._process_with_split(workbook, Path(file_path))
            else:
                self._emit_log("按模式2处理，不进行拆分...", "info")
                self._process_workbook_inplace(workbook)
                output_path = base_name.with_name(f"{base_name.name}_converted").with_suffix(".xlsx")
                workbook.save(output_path)
                output_files = [output_path]
            self._emit_log("图片已批量插入H列，请在Excel中手动设置图片为'嵌入单元格'。", "warning")
            return output_files
        finally:
            self._cleanup()

    # ------------------------------------------------------------------
    # Workbook helpers
    # ------------------------------------------------------------------
    def _load_workbook(self, file_path: str):
        self._emit_log("加载工作簿...", "info")
        try:
            return load_workbook(file_path, data_only=True)
        except Exception:
            self._emit_log("标准模式加载失败，尝试兼容模式...", "warning")
            return load_workbook(file_path, keep_vba=False)

    def _process_workbook_inplace(self, workbook) -> None:
        staff_db = load_staff_database()
        sheets = workbook.worksheets
        total_sheets = len(sheets)
        for index, sheet in enumerate(sheets, start=1):
            self._emit_progress(index, total_sheets, f"处理工作表 {sheet.title}")
            self._process_sheet(sheet, staff_db)

    def _process_sheet(self, sheet, staff_db: Dict[str, str]) -> None:
        h_column = 8
        g_column = 7
        i_column = 9
        t_column = 20
        v_column = 22

        self._set_cell_alignment(sheet)
        self._set_row_heights(sheet, 140)
        sheet.column_dimensions[get_column_letter(h_column)].width = 42

        self._emit_log("进行内容检测和人员匹配...", "info")
        content_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        actor_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        staff_match_count = 0
        content_check_count = 0
        actor_check_count = 0

        for row in range(2, sheet.max_row + 1):
            i_cell = sheet.cell(row=row, column=i_column)
            if i_cell.value and check_content_length(i_cell.value) < 100:
                i_cell.fill = content_fill
                content_check_count += 1

            g_cell = sheet.cell(row=row, column=g_column)
            if g_cell.value and has_invalid_actor_delimiter(g_cell.value):
                g_cell.fill = actor_fill
                actor_check_count += 1

            t_cell = sheet.cell(row=row, column=t_column)
            v_cell = sheet.cell(row=row, column=v_column)
            matched_id = match_staff_id(t_cell.value, staff_db)
            if matched_id:
                v_cell.value = matched_id
                staff_match_count += 1

        self._emit_log(
            f"内容检测完成: I列标红{content_check_count}个, G列标红{actor_check_count}个, 人员匹配{staff_match_count}个",
            "success",
        )

        urls: List[str] = []
        url_rows: List[int] = []
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=h_column).value
            if cell_value and is_valid_url(str(cell_value)):
                urls.append(str(cell_value).strip())
                url_rows.append(row)

        if not urls:
            self._emit_log("未找到有效的URL", "warning")
            return

        self._emit_log(f"找到 {len(urls)} 个URL需要处理", "info")
        download_results = image_fetcher.download_images_concurrently(urls)

        target_width = int(42 * 8)
        target_height = int(140 * 1.33)
        image_paths: Dict[int, Path] = {}

        for idx, row in enumerate(url_rows, start=1):
            self._emit_progress(idx, len(url_rows), "处理图片")
            image_data = download_results.get(idx - 1)
            if not image_data:
                self._emit_log(f"无法下载图片: {urls[idx - 1]}", "error")
                continue
            resized = image_fetcher.resize_image(image_data, target_width, target_height)
            if not resized:
                continue
            path = image_fetcher.save_temp_image(resized, f"{row}", TEMP_DIR)
            self._temp_files.append(path)
            image_paths[row] = path

        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=h_column).value = None

        for row in url_rows:
            temp_path = image_paths.get(row)
            if temp_path and temp_path.exists():
                image = OpenpyxlImage(str(temp_path))
                image.width = target_width
                image.height = target_height
                sheet.add_image(image, f"{get_column_letter(h_column)}{row}")
                cell = sheet.cell(row=row, column=h_column)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        self._emit_log(
            f"工作表 {sheet.title} 处理完成: 成功 {len(image_paths)}/{len(url_rows)}", "success"
        )

    def _process_with_split(self, workbook, original_path: Path) -> List[Path]:
        sheet = workbook.active
        if sheet is None:
            return []

        staff_db = load_staff_database()
        header = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
        data_rows = [
            [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]
            for row in range(2, sheet.max_row + 1)
        ]

        urls: List[str] = []
        url_index_map: List[int] = []
        for row_index, row_data in enumerate(data_rows):
            cell_value = row_data[7 - 1]  # H 列
            if cell_value and is_valid_url(str(cell_value)):
                urls.append(str(cell_value).strip())
                url_index_map.append(row_index)

        download_results = image_fetcher.download_images_concurrently(urls)
        target_width = int(42 * 8)
        target_height = int(140 * 1.33)
        resized_cache: Dict[int, Path] = {}
        for idx, image_data in download_results.items():
            if not image_data:
                continue
            resized = image_fetcher.resize_image(image_data, target_width, target_height)
            if resized:
                row_index = url_index_map[idx]
                path = image_fetcher.save_temp_image(resized, f"split_{row_index}", TEMP_DIR)
                self._temp_files.append(path)
                resized_cache[row_index] = path

        part_size = 50
        total = len(data_rows)
        output_files: List[Path] = []
        for part_index, start in enumerate(range(0, total, part_size), start=1):
            self._emit_progress(part_index, (total + part_size - 1) // part_size, "拆分文件")
            end = min(start + part_size, total)
            chunk = data_rows[start:end]
            wb = Workbook()
            ws = wb.active or wb.create_sheet()
            ws.title = sheet.title
            self._copy_sheet_formatting(sheet, ws)

            for col in range(1, sheet.max_column + 1):
                ws.cell(row=1, column=col, value=header[col - 1])
                self._copy_cell_style(sheet.cell(row=1, column=col), ws.cell(row=1, column=col))

            for row_offset, row_data in enumerate(chunk, start=2):
                for col in range(1, sheet.max_column + 1):
                    ws.cell(row=row_offset, column=col, value=row_data[col - 1])

            self._set_cell_alignment(ws)
            self._set_row_heights(ws, 140)
            ws.column_dimensions[get_column_letter(8)].width = 42

            content_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            actor_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
            staff_match_count = 0
            content_check_count = 0
            actor_check_count = 0

            for row in range(2, ws.max_row + 1):
                i_cell = ws.cell(row=row, column=9)
                if i_cell.value and check_content_length(i_cell.value) < 100:
                    i_cell.fill = content_fill
                    content_check_count += 1

                g_cell = ws.cell(row=row, column=7)
                if g_cell.value and has_invalid_actor_delimiter(g_cell.value):
                    g_cell.fill = actor_fill
                    actor_check_count += 1

                t_cell = ws.cell(row=row, column=20)
                v_cell = ws.cell(row=row, column=22)
                matched_id = match_staff_id(t_cell.value, staff_db)
                if matched_id:
                    v_cell.value = matched_id
                    staff_match_count += 1

            self._emit_log(
                f"内容检测完成: I列标红{content_check_count}个, G列标红{actor_check_count}个, 人员匹配{staff_match_count}个",
                "success",
            )

            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=8).value = None

            for row_offset, row_data in enumerate(chunk):
                source_index = start + row_offset
                image_path = resized_cache.get(source_index)
                if image_path and image_path.exists():
                    img = OpenpyxlImage(str(image_path))
                    img.width = target_width
                    img.height = target_height
                    ws.add_image(img, f"{get_column_letter(8)}{row_offset + 2}")
                    ws.cell(row=row_offset + 2, column=8).alignment = Alignment(horizontal="center", vertical="center")

            output_path = original_path.with_name(
                f"{original_path.stem}_part{part_index}" if total > part_size else f"{original_path.stem}_converted"
            ).with_suffix(".xlsx")
            wb.save(output_path)
            output_files.append(output_path)

        return output_files

    # ------------------------------------------------------------------
    # Formatting helpers
    # ------------------------------------------------------------------
    def _set_cell_alignment(self, sheet) -> None:
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    def _set_row_heights(self, sheet, height: int) -> None:
        for row in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row].height = height

    def _copy_cell_style(self, source_cell, target_cell) -> None:
        if source_cell.font:
            target_cell.font = Font(
                name=source_cell.font.name,
                size=source_cell.font.size,
                bold=source_cell.font.bold,
                italic=source_cell.font.italic,
                color=source_cell.font.color,
            )
        if source_cell.fill:
            target_cell.fill = PatternFill(
                fill_type=source_cell.fill.fill_type,
                start_color=source_cell.fill.start_color,
                end_color=source_cell.fill.end_color,
            )
        if source_cell.border:
            target_cell.border = Border(
                left=source_cell.border.left or Side(style=None),
                right=source_cell.border.right or Side(style=None),
                top=source_cell.border.top or Side(style=None),
                bottom=source_cell.border.bottom or Side(style=None),
            )
        if source_cell.alignment:
            target_cell.alignment = Alignment(
                horizontal=source_cell.alignment.horizontal,
                vertical=source_cell.alignment.vertical,
                wrap_text=source_cell.alignment.wrap_text,
                shrink_to_fit=source_cell.alignment.shrink_to_fit,
            )

    def _copy_sheet_formatting(self, source_sheet, target_sheet) -> None:
        for col in range(1, source_sheet.max_column + 1):
            column_letter = get_column_letter(col)
            if column_letter in source_sheet.column_dimensions:
                target_sheet.column_dimensions[column_letter].width = (
                    source_sheet.column_dimensions[column_letter].width
                )
        for row in range(1, source_sheet.max_row + 1):
            if row in source_sheet.row_dimensions:
                target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

    # ------------------------------------------------------------------
    # Resource management
    # ------------------------------------------------------------------
    def _cleanup(self) -> None:
        for temp_file in self._temp_files:
            try:
                if temp_file.exists():
                    temp_file.unlink()
            except Exception:
                self.logger.warning("删除临时文件失败 %s", temp_file)
        self._temp_files.clear()
        if TEMP_DIR.exists() and not any(TEMP_DIR.iterdir()):
            shutil.rmtree(TEMP_DIR, ignore_errors=True)

